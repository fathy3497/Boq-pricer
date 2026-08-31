import streamlit as st
import pandas as pd
import openpyxl
import json
import io
import requests
from datetime import datetime
import google.generativeai as genai

st.set_page_config(page_title="BOQ Auto Pricer", page_icon="📊", layout="wide")

st.markdown("""
<style>
.main-header { background: #1e3a5f; color: white; padding: 1.2rem 2rem; border-radius: 12px; margin-bottom: 1.5rem; }
.main-header h1 { font-size: 24px; margin: 0; }
.main-header p  { font-size: 13px; opacity: 0.75; margin: 4px 0 0; }
.market-card { background: #f0fdf4; border: 1px solid #86efac; border-radius: 10px; padding: 1rem; margin-bottom: 1rem; }
.market-card h4 { color: #166534; font-size: 13px; margin-bottom: 8px; }
.price-pill { display: inline-block; background: white; border: 1px solid #86efac; border-radius: 20px; padding: 3px 10px; font-size: 12px; margin: 3px; color: #166534; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
  <h1>📊 BOQ Auto Pricer</h1>
  <p>ارفع أي ملف BOQ Excel — أسعار السوق من KAPSARC — سعّر بالذكاء الاصطناعي</p>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ الإعدادات")
    api_key = st.text_input("🔑 Gemini API Key (مجاني)", type="password")
    st.caption("[احصل على مفتاح مجاني ←](https://aistudio.google.com/app/apikey)")

    if api_key:
        try:
            genai.configure(api_key=api_key)
            available = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
            preferred = ['models/gemini-3.6-flash','models/gemini-2.0-flash','models/gemini-1.5-flash','models/gemini-pro']
            auto_model = next((m for m in preferred if m in available), available[0] if available else None)
            selected_model = st.selectbox("🤖 الموديل", available, index=available.index(auto_model) if auto_model in available else 0)
            st.caption(f"✅ {len(available)} موديل متاح")
        except Exception as e:
            st.error(f"خطأ في الـ API Key: {e}")
            selected_model = None
    else:
        selected_model = None

    st.divider()
    region    = st.selectbox("📍 المنطقة", ["Madinah, Saudi Arabia","Riyadh, Saudi Arabia","Jeddah, Saudi Arabia","Cairo, Egypt","Dubai, UAE","Abu Dhabi, UAE","Doha, Qatar"])
    currency  = st.selectbox("💰 العملة", ["SAR","EGP","AED","USD"])
    proj_type = st.selectbox("🏗 نوع المشروع", ["residential","commercial","mixed use","industrial","hospitality"])
    quality   = st.selectbox("⭐ مستوى الجودة", ["standard","high-end","luxury"], index=1)
    st.divider()
    st.caption("v6.0 — BOQ Auto Pricer + KAPSARC")

# ── Fetch market prices from KAPSARC ─────────────────────────────
@st.cache_data(ttl=3600)  # cache for 1 hour
def fetch_kapsarc_prices():
    try:
        url = "https://datasource.kapsarc.org/api/explore/v2.1/catalog/datasets/average-prices-of-some-construction-materials/records"
        params = {"limit": 50, "order_by": "year desc"}
        resp = requests.get(url, params=params, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            records = data.get("results", [])
            prices = {}
            for r in records:
                material = str(r.get("goods") or r.get("material") or r.get("item") or "").strip()
                price    = r.get("price") or r.get("value") or r.get("average_price")
                unit     = r.get("unit") or ""
                year     = r.get("year") or ""
                if material and price:
                    prices[material] = {"price": price, "unit": unit, "year": year}
            return prices, None
        return {}, f"HTTP {resp.status_code}"
    except Exception as e:
        return {}, str(e)

@st.cache_data(ttl=3600)
def fetch_gastat_cci():
    try:
        url = "https://datasource.kapsarc.org/api/explore/v2.1/catalog/datasets/construction-cost-indices-by-sector/records"
        params = {"limit": 10, "order_by": "date desc"}
        resp = requests.get(url, params=params, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            records = data.get("results", [])
            if records:
                latest = records[0]
                return latest, None
        return {}, f"HTTP {resp.status_code}"
    except Exception as e:
        return {}, str(e)

def format_market_context(kapsarc_prices, cci_data):
    """Build a context string for the AI from market data"""
    lines = []
    if kapsarc_prices:
        lines.append("=== Saudi Arabia Market Prices (KAPSARC) ===")
        for mat, info in list(kapsarc_prices.items())[:15]:
            lines.append(f"- {mat}: {info['price']} SAR/{info['unit']} (Year: {info['year']})")
    if cci_data:
        lines.append("\n=== Construction Cost Index (GASTAT) ===")
        for k, v in cci_data.items():
            if v and k not in ['links', 'geo_point_2d']:
                lines.append(f"- {k}: {v}")
    return "\n".join(lines) if lines else ""

# ── BOQ extractor ─────────────────────────────────────────────────
def extract_items(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    all_items = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not ws.max_row or ws.max_row < 3:
            continue
        col_item=col_desc=col_unit=col_qty=-1
        header_row=-1
        for r in range(1, min(25, ws.max_row+1)):
            for c in range(1, min(ws.max_column+1, 20)):
                v = ws.cell(r,c).value
                if not v: continue
                vl = str(v).lower().strip()
                if any(k in vl for k in ['item no','بند','رقم البند','no.']):
                    col_item=c; header_row=r
                if any(k in vl for k in ['description','desc','وصف','particular','work item']):
                    col_desc=c; header_row=r
                if vl in ['unit','وحدة','units']:
                    col_unit=c
                if any(k in vl for k in ['qty','quantity','كمية','quantities']):
                    col_qty=c
            if header_row>0 and col_desc>0 and col_unit>0 and col_qty>0:
                break
        if col_item<0: col_item=1
        if col_desc<0: col_desc=2
        if col_unit<0: col_unit=3
        if col_qty<0:  col_qty=4
        start = header_row+1 if header_row>0 else 3
        rows=[]
        for r in range(start, ws.max_row+1):
            rows.append({'item':ws.cell(r,col_item).value,'desc':ws.cell(r,col_desc).value,
                         'unit':ws.cell(r,col_unit).value,'qty':ws.cell(r,col_qty).value})
        for i,row in enumerate(rows):
            if not row['unit'] or not row['qty']: continue
            try:
                qty_f=float(row['qty'])
                if qty_f<=0: continue
            except: continue
            desc=str(row['desc'] or '').strip()
            item_no=str(row['item'] or '').strip()
            if not desc or not item_no:
                for j in range(i-1,max(-1,i-5),-1):
                    if not desc and rows[j]['desc']: desc=str(rows[j]['desc']).strip()
                    if not item_no and rows[j]['item']: item_no=str(rows[j]['item']).strip()
                    if desc and item_no: break
            lo=(item_no+' '+desc).lower()
            if any(k in lo for k in ['div.','total','summary','carried','subtotal','item no','description','unit price','collection']): continue
            all_items.append({'sheet':sheet_name,'item_no':item_no,'description':desc[:150],
                              'unit':str(row['unit']).strip(),'qty':qty_f,'price':0.0,'total':0.0})
    return all_items

# ── AI Pricing ────────────────────────────────────────────────────
def price_with_gemini(items, api_key, model_name, region, currency, proj_type, quality, market_context=""):
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel(model_name)

    BATCH = 20
    progress = st.progress(0)
    status   = st.empty()
    priced   = 0

    for start in range(0, len(items), BATCH):
        batch = items[start: start+BATCH]
        progress.progress(int(start/len(items)*95))
        status.info(f"⏳ جاري تسعير البنود {start+1}–{min(start+BATCH,len(items))} من {len(items)}...")

        numbered = "\n".join(
            f"{start+j}. {it['description']} (unit: {it['unit']}, qty: {it['qty']})"
            for j,it in enumerate(batch)
        )

        market_section = f"\n\nREFERENCE MARKET DATA (use this to calibrate your prices):\n{market_context}" if market_context else ""

        prompt = f"""You are a senior quantity surveyor in {region}.
Estimate realistic {quality} market unit prices in {currency} for a {proj_type} construction project in {region} (2025/2026).
Include labor, materials, equipment, and contractor overhead.{market_section}

Items:
{numbered}

Return ONLY valid JSON: {{"0": 185, "1": 95, ...}}
No markdown, no explanation."""

        try:
            response = model.generate_content(prompt)
            text = response.text.replace('```json','').replace('```','').strip()
            s=text.find('{'); e=text.rfind('}')
            if s>=0 and e>=0:
                result=json.loads(text[s:e+1])
                for k,v in result.items():
                    idx=int(k)
                    if 0<=idx<len(items) and float(v)>0:
                        items[idx]['price']=float(v)
                        items[idx]['total']=float(v)*items[idx]['qty']
                        priced+=1
        except Exception as ex:
            st.warning(f"⚠ خطأ: {ex}")

    progress.progress(100)
    status.success(f"✅ تم تسعير {priced} بند من أصل {len(items)}")
    return items

# ── Export Excel ──────────────────────────────────────────────────
def to_excel(items, currency):
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "BOQ Priced"
    headers=['رقم البند','الشيت','الوصف','الوحدة','الكمية',f'سعر الوحدة ({currency})',f'الإجمالي ({currency})']
    hf=Font(name='Arial',bold=True,color='FFFFFF',size=11)
    hfill=PatternFill('solid',fgColor='1E3A5F')
    for c,h in enumerate(headers,1):
        cell=ws.cell(1,c,h); cell.font=hf; cell.fill=hfill
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    af=PatternFill('solid',fgColor='EFF6FF')
    nf=Font(name='Arial',size=10,color='1E3A5F',bold=True)
    rf=Font(name='Arial',size=10)
    for i,it in enumerate(items,2):
        ws.cell(i,1,it['item_no']).font=rf; ws.cell(i,2,it['sheet']).font=rf
        ws.cell(i,3,it['description']).font=rf; ws.cell(i,4,it['unit']).font=rf
        ws.cell(i,5,it['qty']).font=nf
        pc=ws.cell(i,6,round(it['price'],2)); pc.font=nf; pc.fill=af; pc.number_format='#,##0.00'
        tc=ws.cell(i,7,f'=F{i}*E{i}'); tc.font=nf; tc.fill=af; tc.number_format='#,##0.00'
    last=len(items)+2
    ws.cell(last,6,'الإجمالي الكلي').font=Font(name='Arial',bold=True,size=11)
    gc=ws.cell(last,7,f'=SUM(G2:G{last-1})')
    gc.font=Font(name='Arial',bold=True,size=12,color='1E3A5F')
    gc.number_format='#,##0.00'; gc.fill=PatternFill('solid',fgColor='DBEAFE')
    for col,w in zip('ABCDEFG',[12,18,52,10,10,20,20]):
        ws.column_dimensions[col].width=w
    ws.freeze_panes='A2'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── Main UI ───────────────────────────────────────────────────────

# Fetch market data automatically
with st.spinner("🔄 جاري جلب أسعار السوق من KAPSARC..."):
    kapsarc_prices, kapsarc_err = fetch_kapsarc_prices()
    cci_data, cci_err = fetch_gastat_cci()

# Show market prices panel
if kapsarc_prices:
    with st.expander("📈 أسعار السوق السعودي — KAPSARC (محدّثة تلقائياً)", expanded=False):
        cols = st.columns(3)
        for i, (mat, info) in enumerate(list(kapsarc_prices.items())[:12]):
            cols[i%3].metric(
                label=mat,
                value=f"{info['price']:,.0f} SAR/{info['unit']}",
                help=f"السنة: {info['year']}"
            )
        st.caption("المصدر: KAPSARC — مركز الملك عبدالله للدراسات البترولية والاقتصادية")
else:
    if kapsarc_err:
        st.info(f"ℹ️ لم يتم جلب أسعار KAPSARC ({kapsarc_err}) — سيعتمد الـ AI على معلوماته الخاصة")

# Build market context for AI
market_context = format_market_context(kapsarc_prices, cci_data)

st.divider()

# File upload
uploaded = st.file_uploader("📂 ارفع ملف BOQ", type=['xlsx','xls'], label_visibility='collapsed')

if uploaded:
    file_bytes = uploaded.read()
    with st.spinner("جاري قراءة الملف..."):
        items = extract_items(file_bytes)
    if not items:
        st.error("⚠ لم يتم اكتشاف أي بنود")
    else:
        st.success(f"✅ تم اكتشاف **{len(items)} بند** من {len(set(i['sheet'] for i in items))} شيت")
        df_preview=pd.DataFrame(items)[['item_no','sheet','description','unit','qty']]
        df_preview.columns=['رقم البند','الشيت','الوصف','الوحدة','الكمية']
        st.dataframe(df_preview,use_container_width=True,height=280)
        st.divider()

        if market_context:
            st.success("✅ الـ AI سيستخدم أسعار KAPSARC كمرجع للتسعير")
        else:
            st.info("ℹ️ الـ AI سيعتمد على معلوماته الخاصة عن السوق")

        if not api_key:
            st.warning("⚠ أدخل Gemini API Key في الشريط الجانبي أولاً")
        elif not selected_model:
            st.warning("⚠ تأكد من صحة الـ API Key")
        else:
            if st.button("✨ ابدأ التسعير التلقائي", use_container_width=True):
                items = price_with_gemini(items, api_key, selected_model, region, currency, proj_type, quality, market_context)
                st.session_state['priced_items'] = items
                st.session_state['currency'] = currency

if 'priced_items' in st.session_state:
    items=st.session_state['priced_items']
    currency=st.session_state['currency']
    st.divider(); st.subheader("📋 النتائج")
    total_priced=sum(1 for it in items if it['price']>0)
    grand_total=sum(it['price']*it['qty'] for it in items)
    c1,c2,c3=st.columns(3)
    c1.metric("إجمالي البنود",len(items))
    c2.metric("بنود مسعّرة",f"{total_priced} / {len(items)}")
    c3.metric(f"الإجمالي ({currency})",f"{grand_total:,.0f}")
    df=pd.DataFrame(items)[['item_no','description','unit','qty','price','total']]
    df.columns=['رقم البند','الوصف','الوحدة','الكمية','سعر الوحدة','الإجمالي']
    st.data_editor(df,use_container_width=True,height=400,
        column_config={
            'سعر الوحدة':st.column_config.NumberColumn(format=f"%.2f {currency}",min_value=0),
            'الإجمالي':st.column_config.NumberColumn(format=f"%.2f {currency}",disabled=True)},
        disabled=['رقم البند','الوصف','الوحدة','الكمية','الإجمالي'])
    excel_buf=to_excel(items,currency)
    st.download_button(label="📥 تحميل Excel",data=excel_buf,
        file_name=f"BOQ_priced_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)
